import argparse
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, List, Optional, Protocol, Tuple

from openpyxl import load_workbook

from rpa_platform.worker.dingtalk_group_handoff import (
    DEFAULT_ELEMENTS_DIR,
    DEFAULT_NORMAL_GROUP_CONFIDENCE,
    DEFAULT_SEARCH_REGION,
    DingtalkGroupHandoffSmokeRunner,
    HandoffElementPaths,
)
from rpa_platform.worker.uia_driver import UiaAutomationDriver


DEFAULT_WORKBOOK = DEFAULT_ELEMENTS_DIR / "需要转交的群.xlsx"
DEFAULT_MEMBER_NAME = "\u5b63\u94b0\u6770"
DEFAULT_MEMBER_ALREADY_IN_REGION = (600, 300, 600, 200)
DEFAULT_MEMBER_ALREADY_IN_CONFIDENCE = 0.70
DEFAULT_ADD_MEMBER_REGION = (1441, 50, 455, 576)
DEFAULT_ADD_MEMBER_CONFIDENCE = 0.75
DEFAULT_SETTINGS_POSITION = (1874, 66)
DEFAULT_CONFIRM_POSITION = (700, 805)
DEFAULT_ADD_MEMBER_POSITION = (1613, 243)
DEFAULT_MEMBER_INPUT_POSITION = (678, 366)
DEFAULT_DINGTALK_READY_POSITION = (793, 963)

ProgressCallback = Callable[[str], None]

STATUS_SUCCESS = "添加成功"
STATUS_MEMBER_ALREADY_IN = "成员已在群内"
STATUS_GROUP_NOT_FOUND = "群不存在"
STATUS_ADD_MEMBER_ENTRY_FAILED = "添加成员入口失败"
STATUS_CONFIRM_NOT_CLICKED = "确认按钮未点击"
STATUS_DINGTALK_WINDOW_NOT_CAPTURED = "钉钉窗口未捕获"


class GroupHandoffBackend(Protocol):
    def handoff_group(self, group_name: str, member_name: str) -> str:
        raise NotImplementedError


class DingtalkWindowNotCaptured(RuntimeError):
    pass


class DingtalkWindowGuard:
    def __init__(
        self,
        window_backend: Any = None,
        activation_backend: Any = None,
        sleep: Any = time.sleep,
        activation_delay_seconds: float = 0.5,
        title_keywords: Tuple[str, ...] = ("钉钉", "DingTalk"),
        progress: Optional[ProgressCallback] = None,
    ) -> None:
        if window_backend is None:
            import pygetwindow  # type: ignore

            window_backend = pygetwindow
        self._window_backend = window_backend
        self._activation_backend = activation_backend
        self._sleep = sleep
        self._activation_delay_seconds = activation_delay_seconds
        self._title_keywords = title_keywords
        self._progress = progress or _noop_progress

    def capture(self) -> str:
        self._emit("捕获钉钉窗口：title_keywords=%s" % "/".join(self._title_keywords))
        windows = self._find_windows()
        if not windows:
            self._emit("未捕获到钉钉窗口，发送 shift+q 唤起")
            self._trigger_dingtalk_activation_shortcut()
            windows = self._find_windows()

        for window in windows:
            title = str(getattr(window, "title", "") or "").strip()
            if title and not _title_matches(title, self._title_keywords):
                continue
            try:
                if bool(getattr(window, "isMinimized", False)) and hasattr(window, "restore"):
                    window.restore()
                if hasattr(window, "activate"):
                    window.activate()
            except Exception as exc:
                raise DingtalkWindowNotCaptured("激活钉钉窗口失败：%s" % _short_error(exc)) from exc
            self._emit("钉钉窗口捕获成功：title=%s" % (title or "钉钉"))
            return title or "钉钉"

        raise DingtalkWindowNotCaptured("未找到标题包含 钉钉/DingTalk 的窗口")

    def _emit(self, message: str) -> None:
        self._progress(message)

    def _find_windows(self) -> List[Any]:
        windows = []
        seen = set()
        for keyword in self._title_keywords:
            try:
                found_windows = self._window_backend.getWindowsWithTitle(keyword)
            except Exception as exc:
                raise DingtalkWindowNotCaptured("查询钉钉窗口失败：%s" % _short_error(exc)) from exc
            for window in found_windows:
                marker = id(window)
                if marker in seen:
                    continue
                seen.add(marker)
                windows.append(window)
        return windows

    def _trigger_dingtalk_activation_shortcut(self) -> None:
        activation_backend = self._activation_backend
        if activation_backend is None:
            import pyautogui  # type: ignore

            activation_backend = pyautogui
        try:
            activation_backend.hotkey("shift", "q")
            self._sleep(self._activation_delay_seconds)
        except Exception as exc:
            raise DingtalkWindowNotCaptured("触发 shift+q 激活钉钉失败：%s" % _short_error(exc)) from exc


@dataclass(frozen=True)
class BatchOptions:
    workbook: Path = DEFAULT_WORKBOOK
    sheet: str = "Sheet1"
    member_name: str = DEFAULT_MEMBER_NAME
    start_row: int = 2
    limit: Optional[int] = None
    dry_run: bool = False
    skip_completed: bool = False
    save_every: int = 1


@dataclass(frozen=True)
class BatchRowResult:
    row: int
    group_name: str
    status: str


@dataclass(frozen=True)
class BatchResult:
    processed_count: int = 0
    failed_count: int = 0
    skipped_count: int = 0
    planned_groups: List[str] = field(default_factory=list)
    rows: List[BatchRowResult] = field(default_factory=list)


class DingtalkGroupHandoffBatchRunner:
    def __init__(
        self,
        backend: GroupHandoffBackend,
        progress: Optional[ProgressCallback] = None,
    ) -> None:
        self._backend = backend
        self._progress = progress or _noop_progress

    def run(self, options: BatchOptions) -> BatchResult:
        workbook_path = Path(options.workbook)
        self._emit(
            "打开 workbook=%s sheet=%s start_row=%s limit=%s skip_completed=%s dry_run=%s"
            % (
                workbook_path,
                options.sheet,
                options.start_row,
                options.limit,
                options.skip_completed,
                options.dry_run,
            )
        )
        workbook = load_workbook(workbook_path)
        sheet = workbook[options.sheet]

        processed = 0
        failed = 0
        skipped = 0
        planned_groups: List[str] = []
        rows: List[BatchRowResult] = []
        changed_since_save = 0

        for row_number in range(int(options.start_row), sheet.max_row + 1):
            group_name = _cell_text(sheet.cell(row=row_number, column=1).value)
            if not group_name:
                continue
            status_cell = sheet.cell(row=row_number, column=2)
            if options.skip_completed and _cell_text(status_cell.value):
                self._emit(
                    "跳过已完成 row=%s group=%s status=%s"
                    % (row_number, group_name, _cell_text(status_cell.value))
                )
                skipped += 1
                continue
            if options.limit is not None and (processed + len(planned_groups)) >= options.limit:
                self._emit("达到 limit=%s，停止继续处理" % options.limit)
                break

            self._emit("开始处理 row=%s group=%s" % (row_number, group_name))
            if options.dry_run:
                self._emit("预览模式 row=%s group=%s，不执行 GUI、不写状态" % (row_number, group_name))
                planned_groups.append(group_name)
                continue

            try:
                status = self._backend.handoff_group(group_name, options.member_name)
            except Exception as exc:
                self._emit(
                    "异常/失败收口 row=%s group=%s error=%s"
                    % (row_number, group_name, _short_error(exc))
                )
                _try_close_dialogs(self._backend)
                status = "异常：%s" % _short_error(exc)

            status_cell.value = status
            rows.append(BatchRowResult(row=row_number, group_name=group_name, status=status))
            self._emit(
                "每行最终状态 row=%s group=%s status=%s"
                % (row_number, group_name, status)
            )
            processed += 1
            if _is_failure_status(status):
                failed += 1
            changed_since_save += 1
            if changed_since_save >= max(int(options.save_every), 1):
                workbook.save(workbook_path)
                self._emit("保存 workbook=%s" % workbook_path)
                changed_since_save = 0

        if changed_since_save:
            workbook.save(workbook_path)
            self._emit("保存 workbook=%s" % workbook_path)

        return BatchResult(
            processed_count=processed,
            failed_count=failed,
            skipped_count=skipped,
            planned_groups=planned_groups,
            rows=rows,
        )

    def _emit(self, message: str) -> None:
        self._progress(message)


class DingtalkGroupHandoffGuiBackend:
    def __init__(
        self,
        elements_dir: Path = DEFAULT_ELEMENTS_DIR,
        uia_driver: Any = None,
        gui_backend: Any = None,
        clipboard_backend: Any = None,
        sleep: Any = time.sleep,
        search_region: Tuple[int, int, int, int] = DEFAULT_SEARCH_REGION,
        normal_group_confidence: float = DEFAULT_NORMAL_GROUP_CONFIDENCE,
        member_already_in_region: Tuple[int, int, int, int] = DEFAULT_MEMBER_ALREADY_IN_REGION,
        member_already_in_confidence: float = DEFAULT_MEMBER_ALREADY_IN_CONFIDENCE,
        add_member_region: Tuple[int, int, int, int] = DEFAULT_ADD_MEMBER_REGION,
        add_member_confidence: float = DEFAULT_ADD_MEMBER_CONFIDENCE,
        settings_position: Tuple[int, int] = DEFAULT_SETTINGS_POSITION,
        confirm_position: Tuple[int, int] = DEFAULT_CONFIRM_POSITION,
        add_member_position: Tuple[int, int] = DEFAULT_ADD_MEMBER_POSITION,
        member_input_position: Tuple[int, int] = DEFAULT_MEMBER_INPUT_POSITION,
        dingtalk_ready_position: Tuple[int, int] = DEFAULT_DINGTALK_READY_POSITION,
        step_delay_seconds: float = 1.0,
        window_guard: Any = None,
        progress: Optional[ProgressCallback] = None,
    ) -> None:
        if gui_backend is None:
            import pyautogui  # type: ignore

            gui_backend = pyautogui
        if clipboard_backend is None:
            import pyperclip  # type: ignore

            clipboard_backend = pyperclip

        self._paths = HandoffElementPaths.from_dir(Path(elements_dir))
        self._add_member_image = Path(elements_dir) / "add_member.png"
        self._member_already_in_image = Path(elements_dir) / "member_already_in.png"
        self._uia_driver = uia_driver or UiaAutomationDriver()
        self._gui = gui_backend
        self._clipboard = clipboard_backend
        self._sleep = sleep
        self._search_region = search_region
        self._normal_group_confidence = normal_group_confidence
        self._member_already_in_region = member_already_in_region
        self._member_already_in_confidence = member_already_in_confidence
        self._add_member_region = add_member_region
        self._add_member_confidence = add_member_confidence
        self._settings_position = settings_position
        self._confirm_position = confirm_position
        self._add_member_position = add_member_position
        self._member_input_position = member_input_position
        self._dingtalk_ready_position = dingtalk_ready_position
        self._step_delay_seconds = step_delay_seconds
        self._window_guard = window_guard or DingtalkWindowGuard()
        self._progress = progress or _noop_progress
        self._smoke_runner = DingtalkGroupHandoffSmokeRunner(
            uia_driver=self._uia_driver,
            gui_backend=self._gui,
            clipboard_backend=self._clipboard,
            sleep=self._sleep,
        )

    def handoff_group(self, group_name: str, member_name: str) -> str:
        self._emit("捕获/唤起钉钉窗口 group=%s" % group_name)
        if not self._capture_dingtalk_window():
            self._emit("钉钉窗口捕获失败 group=%s" % group_name)
            return self._finish_group(group_name, STATUS_DINGTALK_WINDOW_NOT_CAPTURED)

        self._emit("点击钉钉就绪坐标 %s" % _format_position(self._dingtalk_ready_position))
        self._uia_driver.click_position(*self._dingtalk_ready_position)
        self._delay_step()
        self._emit("呼出搜索 ctrl+shift+f")
        self._smoke_runner.open_search(
            self._paths.group_search_input,
            search_open_mode="shortcut",
            click_mode="position",
        )
        self._delay_step()

        self._emit("复制/粘贴群名 group=%s" % group_name)
        self._clipboard.copy(group_name)
        self._delay_step()
        self._gui.hotkey("ctrl", "a")
        self._delay_step()
        self._gui.hotkey("ctrl", "v")
        self._delay_step()

        self._emit("点击群分类")
        self._smoke_runner.click_collected_path(
            self._paths.select_search_type_group,
            click_mode="position",
        )
        self._delay_step()

        self._emit(
            "识别普通群图片 image=%s region=%s confidence=%.2f"
            % (
                self._paths.normal_group_image.name,
                _format_region(self._search_region),
                self._normal_group_confidence,
            )
        )
        if not self._click_image_if_present(
            self._paths.normal_group_image,
            confidence=self._normal_group_confidence,
            region=self._search_region,
        ):
            self._emit("识别普通群图片失败")
            self.close_active_dialogs()
            return self._finish_group(group_name, STATUS_GROUP_NOT_FOUND)
        self._emit("识别普通群图片成功")
        self._delay_step()

        self._emit("点击设置 %s" % _format_position(self._settings_position))
        self._uia_driver.click_position(*self._settings_position)
        self._delay_step()

        if not self._add_member_image.exists():
            self._emit("识别/点击添加成员失败：缺少 image=%s" % self._add_member_image.name)
            self.close_active_dialogs()
            return self._finish_group(group_name, STATUS_ADD_MEMBER_ENTRY_FAILED)
        self._emit(
            "识别/点击添加成员 image=%s region=%s confidence=%.2f"
            % (
                self._add_member_image.name,
                _format_region(self._add_member_region),
                self._add_member_confidence,
            )
        )
        if not self._click_image_if_present(
            self._add_member_image,
            confidence=self._add_member_confidence,
            region=self._add_member_region,
        ):
            self._emit("识别/点击添加成员失败")
            self.close_active_dialogs()
            return self._finish_group(group_name, STATUS_ADD_MEMBER_ENTRY_FAILED)
        self._emit("识别/点击添加成员成功")
        self._delay_step()

        self._emit("点击成员输入框 %s" % _format_position(self._member_input_position))
        self._uia_driver.click_position(*self._member_input_position)
        self._delay_step()
        self._emit("输入成员名 member=%s" % member_name)
        self._clipboard.copy(member_name)
        self._delay_step()
        self._gui.hotkey("ctrl", "a")
        self._delay_step()
        self._gui.hotkey("ctrl", "v")
        self._delay_step()
        self._press("enter")
        self._delay_step()

        self._emit(
            "检测成员已在群内 image=%s region=%s confidence=%.2f"
            % (
                self._member_already_in_image.name,
                _format_region(self._member_already_in_region),
                self._member_already_in_confidence,
            )
        )
        if self._click_image_if_present(
            self._member_already_in_image,
            confidence=self._member_already_in_confidence,
            region=self._member_already_in_region,
            click=False,
        ):
            self._emit("检测成员已在群内：命中")
            self.close_active_dialogs()
            return self._finish_group(group_name, STATUS_MEMBER_ALREADY_IN)
        self._emit("检测成员已在群内：未命中")

        self._emit("点击确认 %s" % _format_position(self._confirm_position))
        try:
            self._uia_driver.click_position(*self._confirm_position)
        except Exception as exc:
            self._emit("点击确认失败：%s" % _short_error(exc))
            self.close_active_dialogs()
            return self._finish_group(group_name, STATUS_CONFIRM_NOT_CLICKED)
        self._delay_step()
        return self._finish_group(group_name, STATUS_SUCCESS)

    def close_active_dialogs(self) -> None:
        self._emit("异常/失败收口：重新捕获钉钉窗口并按 Esc")
        if not self._capture_dingtalk_window():
            self._emit("异常/失败收口：钉钉窗口未捕获，跳过 Esc")
            return
        self._press("esc")
        self._delay_step()
        self._emit("异常/失败收口：已发送 Esc")

    def _finish_group(self, group_name: str, status: str) -> str:
        self._emit("动作最终状态 group=%s status=%s" % (group_name, status))
        return status

    def _emit(self, message: str) -> None:
        self._progress(message)

    def _delay_step(self) -> None:
        self._sleep(self._step_delay_seconds)

    def _capture_dingtalk_window(self) -> bool:
        try:
            self._window_guard.capture()
            return True
        except DingtalkWindowNotCaptured:
            return False

    def _click_image_if_present(
        self,
        image_path: Path,
        confidence: float,
        region: Tuple[int, int, int, int],
        click: bool = True,
    ) -> bool:
        try:
            box = self._gui.locateOnScreen(
                str(image_path),
                confidence=confidence,
                region=region,
            )
        except Exception as exc:
            if exc.__class__.__name__ != "ImageNotFoundException":
                raise
            box = None
        if box is None:
            return False
        if click:
            self._uia_driver.click_position(*_box_center(box))
        return True

    def _press(self, key: str) -> None:
        if hasattr(self._gui, "press"):
            self._gui.press(key)
            return
        self._gui.hotkey(key)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Batch add a DingTalk handoff member from an Excel workbook."
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--sheet", default="Sheet1")
    parser.add_argument("--member-name", default=DEFAULT_MEMBER_NAME)
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-completed", action="store_true")
    parser.add_argument("--save-every", type=int, default=1)
    parser.add_argument("--elements-dir", default=str(DEFAULT_ELEMENTS_DIR))
    parser.add_argument("--pause-before-start", type=float, default=3.0)
    parser.add_argument("--step-delay", type=float, default=1.0)
    args = parser.parse_args(argv)

    options = BatchOptions(
        workbook=Path(args.workbook),
        sheet=args.sheet,
        member_name=args.member_name,
        start_row=args.start_row,
        limit=args.limit,
        dry_run=args.dry_run,
        skip_completed=args.skip_completed,
        save_every=args.save_every,
    )
    if args.dry_run:
        backend: GroupHandoffBackend = _DryRunBackend()
    else:
        _validate_local_assets(Path(args.elements_dir))
        if args.pause_before_start > 0:
            print(
                "Starting in %.1f seconds. Keep DingTalk visible and move the local "
                "mouse out of the RDP window." % args.pause_before_start
            )
            time.sleep(args.pause_before_start)
        window_guard = DingtalkWindowGuard(progress=print)
        try:
            window_title = window_guard.capture()
        except DingtalkWindowNotCaptured as exc:
            print("DingTalk window capture failed: %s" % exc)
            return 2
        print("Captured DingTalk window: %s" % window_title)
        backend = DingtalkGroupHandoffGuiBackend(
            elements_dir=Path(args.elements_dir),
            step_delay_seconds=args.step_delay,
            window_guard=window_guard,
            progress=print,
        )

    result = DingtalkGroupHandoffBatchRunner(backend, progress=print).run(options)
    if args.dry_run:
        for index, group_name in enumerate(result.planned_groups, start=1):
            print("%s. %s" % (index, group_name))
    else:
        print(
            "processed=%s failed=%s skipped=%s workbook=%s"
            % (result.processed_count, result.failed_count, result.skipped_count, options.workbook)
        )
    return 0


class _DryRunBackend:
    def handoff_group(self, group_name: str, member_name: str) -> str:
        raise RuntimeError("dry-run backend should not be called")


def _noop_progress(_message: str) -> None:
    return


def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _short_error(exc: Exception) -> str:
    text = str(exc).strip() or exc.__class__.__name__
    return text[:60]


def _is_failure_status(status: str) -> bool:
    return status.startswith("异常：") or status in {
        STATUS_ADD_MEMBER_ENTRY_FAILED,
        STATUS_CONFIRM_NOT_CLICKED,
        STATUS_DINGTALK_WINDOW_NOT_CAPTURED,
    }


def _try_close_dialogs(backend: Any) -> None:
    close = getattr(backend, "close_active_dialogs", None)
    if close is None:
        return
    try:
        close()
    except Exception:
        return


def _format_position(position: Tuple[int, int]) -> str:
    return "(%s,%s)" % (int(position[0]), int(position[1]))


def _format_region(region: Tuple[int, int, int, int]) -> str:
    return "(%s,%s,%s,%s)" % (
        int(region[0]),
        int(region[1]),
        int(region[2]),
        int(region[3]),
    )


def _validate_local_assets(elements_dir: Path) -> None:
    paths = HandoffElementPaths.from_dir(elements_dir)
    required = [
        paths.group_search_input,
        paths.select_search_type_group,
        paths.normal_group_image,
        paths.add_member_button,
        elements_dir / "add_member.png",
        elements_dir / "member_already_in.png",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing DingTalk handoff assets: %s" % ", ".join(missing))


def _box_center(box: Any) -> Tuple[int, int]:
    left = int(getattr(box, "left", box[0] if isinstance(box, (list, tuple)) else 0))
    top = int(getattr(box, "top", box[1] if isinstance(box, (list, tuple)) else 0))
    width = int(getattr(box, "width", box[2] if isinstance(box, (list, tuple)) else 0))
    height = int(getattr(box, "height", box[3] if isinstance(box, (list, tuple)) else 0))
    return int(left + width / 2), int(top + height / 2)


def _title_matches(title: str, keywords: Tuple[str, ...]) -> bool:
    title_lower = title.lower()
    return any(keyword.lower() in title_lower for keyword in keywords)


if __name__ == "__main__":
    raise SystemExit(main())
