import json
import inspect
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from rpa_platform.worker.dingtalk_group_handoff_batch import (
    BatchOptions,
    DingtalkGroupHandoffGuiBackend,
    DingtalkGroupHandoffBatchRunner,
    DingtalkWindowGuard,
    DingtalkWindowNotCaptured,
    STATUS_ADD_MEMBER_ENTRY_FAILED,
    STATUS_DINGTALK_WINDOW_NOT_CAPTURED,
    STATUS_GROUP_NOT_FOUND,
    STATUS_SUCCESS,
)


class DingtalkGroupHandoffBatchTest(unittest.TestCase):
    def test_batch_runner_reports_row_progress_failures_and_workbook_saves(self):
        self.assertIn(
            "progress",
            inspect.signature(DingtalkGroupHandoffBatchRunner).parameters,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = _write_workbook(
                Path(tmpdir) / "groups.xlsx",
                [
                    ("群名称", "群状态"),
                    ("群A", ""),
                    ("群B", ""),
                ],
            )
            backend = _FakeHandoffBackend(
                {
                    "群A": STATUS_SUCCESS,
                    "群B": RuntimeError("添加成员入口失败"),
                }
            )
            messages = []

            DingtalkGroupHandoffBatchRunner(backend, progress=messages.append).run(
                BatchOptions(workbook=workbook, save_every=1)
            )

            output = "\n".join(messages)
            self.assertIn("开始处理 row=2 group=群A", output)
            self.assertIn("每行最终状态 row=2 group=群A status=添加成功", output)
            self.assertIn("开始处理 row=3 group=群B", output)
            self.assertIn("异常/失败收口 row=3 group=群B error=添加成员入口失败", output)
            self.assertIn("每行最终状态 row=3 group=群B status=异常：添加成员入口失败", output)
            self.assertIn("保存 workbook", output)

    def test_processes_non_empty_groups_writes_status_and_continues_after_error(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = _write_workbook(
                Path(tmpdir) / "groups.xlsx",
                [
                    ("群名称", "群状态"),
                    ("群A", ""),
                    ("", ""),
                    ("群B", ""),
                    ("群C", ""),
                ],
            )
            backend = _FakeHandoffBackend(
                {
                    "群A": "添加成功",
                    "群B": RuntimeError("添加成员入口失败"),
                    "群C": "成员已在群内",
                }
            )

            result = DingtalkGroupHandoffBatchRunner(backend).run(
                BatchOptions(workbook=workbook)
            )

            self.assertEqual(result.processed_count, 3)
            self.assertEqual(result.failed_count, 1)
            self.assertEqual(backend.calls, [("群A", "季钰杰"), ("群B", "季钰杰"), ("群C", "季钰杰")])
            self.assertEqual(_column_values(workbook, "B", 2, 5), ["添加成功", None, "异常：添加成员入口失败", "成员已在群内"])

    def test_skip_completed_and_limit_apply_before_gui_work(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = _write_workbook(
                Path(tmpdir) / "groups.xlsx",
                [
                    ("群名称", "群状态"),
                    ("已完成群", "添加成功"),
                    ("群A", ""),
                    ("群B", ""),
                    ("群C", ""),
                ],
            )
            backend = _FakeHandoffBackend({"群A": "群不存在", "群B": "添加成功", "群C": "添加成功"})

            result = DingtalkGroupHandoffBatchRunner(backend).run(
                BatchOptions(workbook=workbook, skip_completed=True, limit=2)
            )

            self.assertEqual(result.processed_count, 2)
            self.assertEqual(backend.calls, [("群A", "季钰杰"), ("群B", "季钰杰")])
            self.assertEqual(_column_values(workbook, "B", 2, 5), ["添加成功", "群不存在", "添加成功", None])

    def test_dry_run_lists_groups_without_writing_or_clicking(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = _write_workbook(
                Path(tmpdir) / "groups.xlsx",
                [
                    ("群名称", "群状态"),
                    ("群A", ""),
                    ("群B", ""),
                ],
            )
            backend = _FakeHandoffBackend({"群A": "添加成功", "群B": "添加成功"})

            result = DingtalkGroupHandoffBatchRunner(backend).run(
                BatchOptions(workbook=workbook, dry_run=True)
            )

            self.assertEqual(result.processed_count, 0)
            self.assertEqual(result.planned_groups, ["群A", "群B"])
            self.assertEqual(backend.calls, [])
            self.assertEqual(_column_values(workbook, "B", 2, 3), [None, None])

    def test_dev_script_entrypoint_supports_dry_run_limit(self):
        from scripts.dev.run_dingtalk_group_handoff_batch import main

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook = _write_workbook(
                Path(tmpdir) / "groups.xlsx",
                [
                    ("群名称", "群状态"),
                    ("群A", ""),
                    ("群B", ""),
                ],
            )
            output = StringIO()

            with redirect_stdout(output):
                exit_code = main(["--workbook", str(workbook), "--dry-run", "--limit", "1"])

            self.assertEqual(exit_code, 0)
            self.assertIn("1. 群A", output.getvalue())
            self.assertNotIn("群B", output.getvalue())
            self.assertEqual(_column_values(workbook, "B", 2, 3), [None, None])

    def test_gui_backend_returns_add_member_entry_failed_when_add_member_image_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            elements_dir = _write_handoff_files(Path(tmpdir))
            calls = []
            backend = DingtalkGroupHandoffGuiBackend(
                elements_dir=elements_dir,
                uia_driver=_FakeDriver(calls),
                gui_backend=_FakeGui(
                    calls,
                    image_results={
                        "normal_group.png": SimpleNamespace(left=617, top=112, width=42, height=24),
                        "add_member.png": None,
                    },
                ),
                clipboard_backend=_FakeClipboard(calls),
                sleep=lambda _seconds: None,
                window_guard=_FakeWindowGuard(calls),
            )

            status = backend.handoff_group("群A", "季钰杰")

            self.assertEqual(status, STATUS_ADD_MEMBER_ENTRY_FAILED)
            self.assertEqual(calls[0], ("capture_window",))
            self.assertIn(("locate", "add_member.png", 0.75, (1441, 50, 455, 576)), calls)
            self.assertNotIn(("position_click", 700, 805), calls)

    def test_gui_backend_closes_search_overlay_when_group_is_not_found(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            elements_dir = _write_handoff_files(Path(tmpdir))
            calls = []
            backend = DingtalkGroupHandoffGuiBackend(
                elements_dir=elements_dir,
                uia_driver=_FakeDriver(calls),
                gui_backend=_FakeGui(
                    calls,
                    image_results={
                        "normal_group.png": None,
                    },
                ),
                clipboard_backend=_FakeClipboard(calls),
                sleep=lambda _seconds: None,
                window_guard=_FakeWindowGuard(calls),
            )

            status = backend.handoff_group("群A", "季钰杰")

            self.assertEqual(status, STATUS_GROUP_NOT_FOUND)
            self.assertEqual(calls.count(("capture_window",)), 2)
            self.assertEqual(calls.count(("press", "esc")), 1)
            self.assertNotIn(("position_click", 1874, 66), calls)

    def test_gui_backend_clicks_add_member_image_when_detected(self):
        self.assertIn(
            "progress",
            inspect.signature(DingtalkGroupHandoffGuiBackend).parameters,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            elements_dir = _write_handoff_files(Path(tmpdir))
            calls = []
            messages = []
            backend = DingtalkGroupHandoffGuiBackend(
                elements_dir=elements_dir,
                uia_driver=_FakeDriver(calls),
                gui_backend=_FakeGui(
                    calls,
                    image_results={
                        "normal_group.png": SimpleNamespace(left=617, top=112, width=42, height=24),
                        "add_member.png": SimpleNamespace(left=1600, top=230, width=34, height=26),
                        "member_already_in.png": None,
                    },
                ),
                clipboard_backend=_FakeClipboard(calls),
                sleep=lambda seconds: calls.append(("sleep", seconds)),
                window_guard=_FakeWindowGuard(calls),
                progress=messages.append,
            )

            status = backend.handoff_group("群A", "季钰杰")

            self.assertEqual(status, STATUS_SUCCESS)
            self.assertEqual(calls[0], ("capture_window",))
            self.assertIn(
                [
                    ("capture_window",),
                    ("position_click", 793, 963),
                    ("sleep", 1.0),
                    ("hotkey", ("ctrl", "shift", "f")),
                    ("sleep", 1.0),
                    ("clipboard_copy", "群A"),
                ],
                _windows(calls, 6),
            )
            self.assertIn(("locate", "add_member.png", 0.75, (1441, 50, 455, 576)), calls)
            self.assertIn(("position_click", 1617, 243), calls)
            self.assertIn(
                [
                    ("position_click", 1617, 243),
                    ("sleep", 1.0),
                    ("position_click", 678, 366),
                    ("sleep", 1.0),
                    ("clipboard_copy", "季钰杰"),
                    ("sleep", 1.0),
                    ("hotkey", ("ctrl", "a")),
                    ("sleep", 1.0),
                    ("hotkey", ("ctrl", "v")),
                    ("sleep", 1.0),
                    ("press", "enter"),
                    ("sleep", 1.0),
                ],
                _windows(calls, 12),
            )
            output = "\n".join(messages)
            self.assertIn("捕获/唤起钉钉窗口", output)
            self.assertIn("点击钉钉就绪坐标 (793,963)", output)
            self.assertIn("呼出搜索 ctrl+shift+f", output)
            self.assertIn("复制/粘贴群名 group=群A", output)
            self.assertIn("点击群分类", output)
            self.assertIn("识别普通群图片成功", output)
            self.assertIn("点击设置 (1874,66)", output)
            self.assertIn("识别/点击添加成员成功", output)
            self.assertIn("点击成员输入框 (678,366)", output)
            self.assertIn("输入成员名 member=季钰杰", output)
            self.assertIn("检测成员已在群内：未命中", output)
            self.assertIn("点击确认 (700,805)", output)
            self.assertIn("动作最终状态 group=群A status=添加成功", output)

    def test_gui_backend_reports_group_not_found_and_closeout(self):
        self.assertIn(
            "progress",
            inspect.signature(DingtalkGroupHandoffGuiBackend).parameters,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            elements_dir = _write_handoff_files(Path(tmpdir))
            calls = []
            messages = []
            backend = DingtalkGroupHandoffGuiBackend(
                elements_dir=elements_dir,
                uia_driver=_FakeDriver(calls),
                gui_backend=_FakeGui(
                    calls,
                    image_results={
                        "normal_group.png": None,
                    },
                ),
                clipboard_backend=_FakeClipboard(calls),
                sleep=lambda _seconds: None,
                window_guard=_FakeWindowGuard(calls),
                progress=messages.append,
            )

            status = backend.handoff_group("群A", "季钰杰")

            self.assertEqual(status, STATUS_GROUP_NOT_FOUND)
            output = "\n".join(messages)
            self.assertIn("识别普通群图片失败", output)
            self.assertIn("异常/失败收口：重新捕获钉钉窗口并按 Esc", output)
            self.assertIn("动作最终状态 group=群A status=群不存在", output)

    def test_gui_backend_blocks_when_dingtalk_window_is_not_captured(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            elements_dir = _write_handoff_files(Path(tmpdir))
            calls = []
            backend = DingtalkGroupHandoffGuiBackend(
                elements_dir=elements_dir,
                uia_driver=_FakeDriver(calls),
                gui_backend=_FakeGui(calls, image_results={}),
                clipboard_backend=_FakeClipboard(calls),
                sleep=lambda _seconds: None,
                window_guard=_FakeWindowGuard(calls, captured=False),
            )

            status = backend.handoff_group("群A", "季钰杰")

            self.assertEqual(status, STATUS_DINGTALK_WINDOW_NOT_CAPTURED)
            self.assertEqual(calls, [("capture_window",)])

    def test_window_guard_restores_and_activates_matching_dingtalk_window(self):
        calls = []
        window = _FakeWindow("钉钉", calls, minimized=True)
        guard = DingtalkWindowGuard(
            window_backend=_FakeWindowBackend([window]),
            activation_backend=_FakeActivationBackend(calls),
            sleep=lambda _seconds: None,
        )

        title = guard.capture()

        self.assertEqual(title, "钉钉")
        self.assertEqual(calls, [("restore", "钉钉"), ("activate", "钉钉")])

    def test_window_guard_uses_shift_q_then_retries_when_window_is_missing(self):
        calls = []
        windows = []
        activation_backend = _FakeActivationBackend(
            calls,
            on_hotkey=lambda: windows.append(_FakeWindow("钉钉", calls)),
        )
        guard = DingtalkWindowGuard(
            window_backend=_FakeWindowBackend(windows),
            activation_backend=activation_backend,
            sleep=lambda seconds: calls.append(("sleep", seconds)),
            activation_delay_seconds=0.1,
        )

        title = guard.capture()

        self.assertEqual(title, "钉钉")
        self.assertEqual(
            calls,
            [
                ("hotkey", ("shift", "q")),
                ("sleep", 0.1),
                ("activate", "钉钉"),
            ],
        )

    def test_window_guard_raises_when_dingtalk_window_is_missing(self):
        calls = []
        guard = DingtalkWindowGuard(
            window_backend=_FakeWindowBackend([]),
            activation_backend=_FakeActivationBackend(calls),
            sleep=lambda _seconds: None,
        )

        with self.assertRaises(DingtalkWindowNotCaptured):
            guard.capture()
        self.assertEqual(calls, [("hotkey", ("shift", "q"))])


class _FakeHandoffBackend:
    def __init__(self, outcomes):
        self.outcomes = outcomes
        self.calls = []

    def handoff_group(self, group_name, member_name):
        self.calls.append((group_name, member_name))
        outcome = self.outcomes[group_name]
        if isinstance(outcome, Exception):
            raise outcome
        return outcome


class _FakeDriver:
    def __init__(self, calls):
        self.calls = calls

    def click_position(self, x, y):
        self.calls.append(("position_click", x, y))


class _FakeGui:
    def __init__(self, calls, image_results):
        self.calls = calls
        self.image_results = image_results

    def hotkey(self, *keys):
        self.calls.append(("hotkey", tuple(keys)))

    def press(self, key):
        self.calls.append(("press", key))

    def locateOnScreen(self, image, confidence, region):
        image_name = Path(image).name
        self.calls.append(("locate", image_name, confidence, region))
        return self.image_results.get(image_name)


class _FakeClipboard:
    def __init__(self, calls):
        self.calls = calls

    def copy(self, value):
        self.calls.append(("clipboard_copy", value))


class _FakeWindowGuard:
    def __init__(self, calls, captured=True):
        self.calls = calls
        self.captured = captured

    def capture(self):
        self.calls.append(("capture_window",))
        if not self.captured:
            raise DingtalkWindowNotCaptured("missing")
        return "钉钉"


class _FakeWindowBackend:
    def __init__(self, windows):
        self.windows = windows

    def getWindowsWithTitle(self, title):
        title_lower = title.lower()
        return [window for window in self.windows if title_lower in window.title.lower()]


class _FakeActivationBackend:
    def __init__(self, calls, on_hotkey=None):
        self.calls = calls
        self.on_hotkey = on_hotkey

    def hotkey(self, *keys):
        self.calls.append(("hotkey", tuple(keys)))
        if self.on_hotkey is not None:
            self.on_hotkey()


class _FakeWindow:
    def __init__(self, title, calls, minimized=False):
        self.title = title
        self.calls = calls
        self.isMinimized = minimized

    def restore(self):
        self.calls.append(("restore", self.title))
        self.isMinimized = False

    def activate(self):
        self.calls.append(("activate", self.title))


def _write_workbook(path: Path, rows) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def _write_handoff_files(root: Path) -> Path:
    elements_dir = root / "dingtalk_group_handoff"
    elements_dir.mkdir()
    for filename in ("normal_group.png", "add_member.png", "member_already_in.png"):
        (elements_dir / filename).write_bytes(b"fake-image")

    files = {
        "group_search_input.json": {
            "target": _target("search_input", "EditControl"),
            "fallback_position": {"x": 650, "y": 20},
        },
        "select_search_type_group.json": {
            "target": _target("select_group", "TextControl"),
            "fallback_position": {"x": 610, "y": 80},
        },
        "group_settings_button.json": {
            "target": _target("settings_button", "ButtonControl"),
            "fallback_position": {"x": 1874, "y": 66},
        },
        "add_member_button.json": {
            "target": _target("add_member_button", "ButtonControl"),
            "fallback_position": {"x": 1613, "y": 243},
        },
    }
    for filename, payload in files.items():
        (elements_dir / filename).write_text(
            json.dumps(payload, ensure_ascii=False),
            encoding="utf-8",
        )
    return elements_dir


def _target(automation_id: str, control_type: str) -> dict:
    return {
        "type": "uia",
        "window_title": "钉钉",
        "control_type": control_type,
        "name": automation_id,
        "automation_id": automation_id,
        "class_name": "",
        "bounding_rect_hint": [1, 2, 3, 4],
    }


def _column_values(path: Path, column: str, start: int, end: int):
    workbook = load_workbook(path)
    sheet = workbook["Sheet1"]
    return [sheet[f"{column}{row}"].value for row in range(start, end + 1)]


def _windows(values, size):
    return [values[index : index + size] for index in range(0, len(values) - size + 1)]


if __name__ == "__main__":
    unittest.main()
